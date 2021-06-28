import FwGuide
import MySQLdb
from openpyxl import Workbook
from openpyxl import load_workbook


def init_class():
    workbook = load_workbook("./FW_Guide_List.xlsx")
    worksheet = workbook.active
    rows = worksheet.max_row
    columns = worksheet.max_column
    fw_list = []
    for row in range(rows - 3):
        if worksheet.cell(row + 3, 1).value is None:
            break
        guide = FwGuide.FwGuide(worksheet.cell(row + 3, 1).value, worksheet.cell(row + 3, 2).value,
                                FwGuide.PID(worksheet.cell(row + 3, 3).value,
                                            worksheet.cell(row + 3, 4).value,
                                            worksheet.cell(row + 3, 5).value),
                                FwGuide.PCB(worksheet.cell(row + 3, 6).value,
                                            worksheet.cell(row + 3, 7).value,
                                            worksheet.cell(row + 3, 8).value),
                                FwGuide.FilePath(worksheet.cell(row + 3, 9).value,
                                                 worksheet.cell(row + 3, 10).value,
                                                 worksheet.cell(row + 3, 11).value),
                                FwGuide.BurnWay(worksheet.cell(row + 3, 12).value,
                                                worksheet.cell(row + 3, 13).value,
                                                worksheet.cell(row + 3, 14).value))

        fw_list.insert(row, guide)
    print(fw_list)


def init_fw():
    workbook = load_workbook("./fw.xlsx")
    worksheet = workbook.active
    rows = worksheet.max_row
    columns = worksheet.max_column
    path_list = []
    for row in range(rows - 1):
        fw_file_path = FwGuide.FWFilePath(row, worksheet.cell(row + 2, 1).value, worksheet.cell(row + 2, 2).value,
                                          worksheet.cell(row + 2, 4).value)
        path_list.insert(row, fw_file_path)
    return path_list


def upload2DB():
    db = MySQLdb.connect("121.196.53.11", "root", "yh0769", "fw_db", charset='utf8')
    cursor = db.cursor()
    fw_list = init_fw()

    for fw in fw_list:
        cursor.execute("insert into FW_PATH (id, fw_page, model, filepath) values('%s', '%s', '%s', '%s')" %
                       (fw.id + 1, fw.fw_page, fw.model, fw.file_path))

    db.commit()


def print_hi(name):
    print(f'Hi, {name}')


if __name__ == '__main__':
    init_class()
    # upload2DB()
